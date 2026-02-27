#!/usr/bin/env python3
"""
从Word文档提取表格数据并保存为Excel文件
支持保留表格格式和结构
"""

import argparse
import sys
import os
from pathlib import Path
from typing import List, Dict, Any
import pandas as pd
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def extract_tables_from_docx(docx_path: str) -> List[List[List[str]]]:
    """
    从Word文档提取所有表格

    Args:
        docx_path: Word文档路径

    Returns:
        表格数据列表，每个表格是二维列表
    """
    try:
        doc = Document(docx_path)
        tables_data = []

        for table in doc.tables:
            table_data = []
            for row in table.rows:
                row_data = []
                for cell in row.cells:
                    # 提取单元格文本，去除多余空白
                    text = cell.text.strip()
                    row_data.append(text)
                table_data.append(row_data)
            tables_data.append(table_data)

        return tables_data
    except Exception as e:
        print(f"读取Word文档失败: {e}")
        return []


def create_excel_with_tables(tables_data: List[List[List[str]]],
                            output_path: str,
                            sheet_name: str = "Tables",
                            preserve_format: bool = False) -> bool:
    """
    将表格数据保存到Excel文件

    Args:
        tables_data: 表格数据列表
        output_path: 输出Excel文件路径
        sheet_name: 工作表名称
        preserve_format: 是否保留格式

    Returns:
        是否成功
    """
    try:
        if not tables_data:
            print("警告: 未找到表格数据")
            return False

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        current_row = 1

        for table_idx, table in enumerate(tables_data):
            if table_idx > 0:
                # 表格之间添加空行
                current_row += 2

            # 添加表格标题
            if len(tables_data) > 1:
                title_cell = ws.cell(row=current_row, column=1, value=f"表格 {table_idx + 1}")
                if preserve_format:
                    title_cell.font = Font(bold=True, size=14)
                    title_cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                current_row += 1

            # 写入表格数据
            for row_idx, row in enumerate(table):
                for col_idx, cell_value in enumerate(row):
                    cell = ws.cell(row=current_row + row_idx,
                                  column=col_idx + 1,
                                  value=cell_value)

                    if preserve_format:
                        # 应用基本格式
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                        # 设置边框
                        thin_border = Border(left=Side(style='thin'),
                                           right=Side(style='thin'),
                                           top=Side(style='thin'),
                                           bottom=Side(style='thin'))
                        cell.border = thin_border

                        # 表头加粗
                        if row_idx == 0:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

            # 调整列宽
            if table:
                for col_idx in range(len(table[0])):
                    max_length = 0
                    for row in table:
                        if col_idx < len(row):
                            cell_length = len(str(row[col_idx]))
                            if cell_length > max_length:
                                max_length = cell_length

                    # 设置列宽，最小10，最大50
                    column_width = min(max(max_length + 2, 10), 50)
                    ws.column_dimensions[get_column_letter(col_idx + 1)].width = column_width

            current_row += len(table)

        # 保存文件
        wb.save(output_path)
        print(f"Excel文件已保存: {output_path}")
        print(f"提取了 {len(tables_data)} 个表格")
        return True

    except Exception as e:
        print(f"保存Excel文件失败: {e}")
        return False


def main():
    parser = argparse.ArgumentParser(description='从Word文档提取表格数据到Excel')
    parser.add_argument('input', help='输入Word文档路径 (.docx)')
    parser.add_argument('--output', '-o', default='output.xlsx', help='输出Excel文件路径 (默认: output.xlsx)')
    parser.add_argument('--sheet-name', '-s', default='Tables', help='Excel工作表名称 (默认: Tables)')
    parser.add_argument('--preserve-format', '-f', action='store_true', help='保留表格格式')
    parser.add_argument('--verbose', '-v', action='store_true', help='显示详细信息')

    args = parser.parse_args()

    # 检查输入文件
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"错误: 输入文件不存在: {args.input}")
        sys.exit(1)

    if input_path.suffix.lower() != '.docx':
        print(f"警告: 文件扩展名不是 .docx: {args.input}")

    # 提取表格数据
    if args.verbose:
        print(f"正在读取文档: {args.input}")

    tables_data = extract_tables_from_docx(str(input_path))

    if args.verbose:
        print(f"找到 {len(tables_data)} 个表格")
        for idx, table in enumerate(tables_data):
            print(f"  表格 {idx + 1}: {len(table)} 行 × {len(table[0]) if table else 0} 列")

    # 保存到Excel
    success = create_excel_with_tables(
        tables_data,
        args.output,
        args.sheet_name,
        args.preserve_format
    )

    if success:
        print("处理完成!")
        sys.exit(0)
    else:
        print("处理失败!")
        sys.exit(1)


if __name__ == "__main__":
    main()